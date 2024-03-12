import json
import os

from openpyxl import load_workbook
from openpyxl import Workbook

from whatsapp_api_client_python import API
from main.forms import UserInfoForm
from main.models import User
from .serializers import VisitSerializer
from rest_framework import generics
from rest_framework.response import Response

from .consts import TELE_NUM, API_TOKEN_INSTANCE, ID_INSTANCE


class VisitCard(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = VisitSerializer

    def perform_create(self, serializer):
        instance = serializer.save()

        file_path = 'fixtures/main/user_info.json'
        try:
            if not os.path.exists(file_path):
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump([], file)
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
        except (json.JSONDecodeError, FileNotFoundError):
            data = []

        serialized_data = VisitSerializer(instance).data
        data.append(serialized_data)

        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(data, file, ensure_ascii=False)

    def whatsapp_sender(self, form_data):
        try:
            if "IdInstance" in locals() or "ApiTokenInstance" in locals():
                return

            greenAPI = API.GreenAPI(ID_INSTANCE, API_TOKEN_INSTANCE)
            payload = f'Имя: {form_data["name"]}\n' \
                      f'Фамилия: {form_data["surname"]}\n' \
                      f'Город: {form_data["city"]}\n' \
                      f'Телефон: {form_data["phone"]}\r\n' \
                      f'Товары: {form_data["products"]}' \

            greenAPI.sending.sendMessage(f"{TELE_NUM}@c.us", payload)

        except Exception as e:
            print(f'An error occurred: {str(e)}')

    def export_form_data_to_xlsx(self, form_data):
        file_path = 'fixtures/main/user_info.xlsx'

        try:
            load = load_workbook(file_path)
        except FileNotFoundError:
            load = Workbook()
            load.create_sheet('users_info')

        sheet = load['users_info']

        data = [
            form_data['name'],
            form_data['surname'],
            form_data['city'],
            form_data['phone'],
            form_data['products']
        ]

        sheet.append(data)
        load.save(file_path)

    def post(self, request):
        form = UserInfoForm(request.POST or None)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        self.whatsapp_sender(serializer.data)
        self.export_form_data_to_xlsx(serializer.data)
        return Response({'message': request.data})

    def get(self, request):
        queryset = User.objects.all()
        serializer = VisitSerializer(queryset, many=True)
        return Response(serializer.data)
